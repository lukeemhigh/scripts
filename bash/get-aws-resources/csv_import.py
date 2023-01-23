#!/usr/bin/env python3
#
# Imports .csv files into an Excel workbook
#
# Author: Luca Giugliardi
# Email: luca.giugliardi@gmail.com

import argparse
import pandas as pd
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import os
import datetime

parser = argparse.ArgumentParser(description='Imports csv into an Excel workbook.')

parser.add_argument("-d", "--directory", help="Full path to the directory in which this script will operate", required=True)

args = parser.parse_args()

date = datetime.date.today()

path = f'{args.directory}'

os.chdir(f'{path}/data')
file_list = os.listdir()

df_list = []

for file in file_list:
    df = pd.read_csv(file, names=["Region", "Service", "Resource", "Key", "Value"])
    df_list.append(df)

df_tot = pd.concat(df_list, ignore_index=True)
os.chdir(f'{path}/output')

def bg_header(x):
    return "background-color: orange"

file_name = f'res-it_aws_resources_{date}.xlsx'
df_tot.style.applymap_index(bg_header, axis=1).to_excel(file_name, index=False)

wb = openpyxl.load_workbook(file_name)
worksheet = wb.active

for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        if cell.coordinate in worksheet.merged_cells: # not check merge_cells
            continue
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = max_length + 0.5
    worksheet.column_dimensions[str(column)].width = adjusted_width

wb.save(file_name)